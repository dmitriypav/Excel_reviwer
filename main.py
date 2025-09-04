from openpyxl import load_workbook

# Load Excel file
file_path = "Data/file_name.xlsx"
wb = load_workbook(file_path)
ws = wb["Sheet1"]


# Add new column
result_col = ws.max_column + 1
ws.cell(row=1, column=result_col, value='Result')

# Extract value from cell B2
value_b2 = ws["B2"].value
# Verification Address
if value_b2 != '':
    ws.cell(row=2, column=result_col, value='OK')
else:
    ws.cell(row=2, column=result_col, value=f'FAIL {value_b2}')

# Value B3
value_b3 = ws["B3"].value
# Value B4
value_b4 = ws["B4"].value

if value_b3 == 'FortiNet' and value_b4 == 'FortiGate 70D':
    ws.cell(row=3, column=result_col, value=f'OK \nFirewall model {value_b4}')
else:
    ws.cell(row=3, column=result_col, value=f'FALSE \nFirewall model {value_b4}')

# Save to file
output_path = 'Data/file_name-verified.xlsx'
wb.save(output_path)

print('File saved:', output_path)